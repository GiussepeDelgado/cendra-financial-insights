from io import BytesIO

import pandas as pd
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.utils import get_column_letter


def generate_excel_report(
    df: pd.DataFrame,
    kpis: dict,
    commercial_kpis: dict,
    cendra_score: dict,
    score_classification: dict,
    concentration_risk: dict,
    insights: list[dict],
) -> bytes:

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:

        # =========================
        # RESUMEN EJECUTIVO
        # =========================

        summary_rows = [
            ["CENDRA - REPORTE EJECUTIVO", ""],
            ["", ""],

            ["INDICADORES FINANCIEROS", ""],
            ["Ventas", kpis["sales"]],
            ["Gastos", kpis["expenses"]],
            ["Utilidad", kpis["profit"]],
            ["Margen", kpis["margin"] / 100],

            ["", ""],

            ["INDICADORES COMERCIALES", ""],
            [
                "Ticket promedio",
                commercial_kpis["average_ticket"],
            ],
            [
                "Número de ventas",
                commercial_kpis["number_of_sales"],
            ],
            [
                "Producto líder",
                commercial_kpis["top_product"],
            ],
            [
                "Cliente principal",
                commercial_kpis["top_customer"],
            ],

            ["", ""],

            ["CENDRA SCORE", ""],
            [
                "Puntuación total",
                cendra_score["total"],
            ],
            [
                "Clasificación",
                score_classification["status"],
            ],
            [
                "Cobertura del análisis",
                cendra_score["coverage"] / 100,
            ],
            [
                "Rentabilidad",
                cendra_score["profitability"],
            ],
            [
                "Crecimiento",
                cendra_score["growth"],
            ],
            [
                "Control de gastos",
                cendra_score["expense_control"],
            ],
            [
                "Estabilidad",
                cendra_score["stability"],
            ],

            ["", ""],

            ["RIESGO DE CONCENTRACIÓN", ""],
            [
                "Cliente principal",
                concentration_risk["top_customer"],
            ],
            [
                "Participación cliente",
                concentration_risk["customer_share"] / 100,
            ],
            [
                "Producto principal",
                concentration_risk["top_product"],
            ],
            [
                "Participación producto",
                concentration_risk["product_share"] / 100,
            ],
        ]

        summary_df = pd.DataFrame(
            summary_rows,
            columns=[
                "Indicador",
                "Valor",
            ],
        )

        summary_df.to_excel(
            writer,
            sheet_name="Resumen",
            index=False,
        )

        # =========================
        # SEÑALES
        # =========================

        if insights:

            insights_df = pd.DataFrame([
                {
                    "Tipo": insight["type"],
                    "Título": insight["title"],
                    "Mensaje": insight["message"],
                }
                for insight in insights
            ])

        else:

            insights_df = pd.DataFrame([
                {
                    "Tipo": "info",
                    "Título": "Sin señales",
                    "Mensaje": (
                        "No se detectaron señales "
                        "relevantes en el período."
                    ),
                }
            ])

        insights_df.to_excel(
            writer,
            sheet_name="Señales",
            index=False,
        )

        # =========================
        # TRANSACCIONES
        # =========================

        export_df = df.copy()

        export_df["fecha"] = (
            export_df["fecha"].dt.date
        )

        export_df.to_excel(
            writer,
            sheet_name="Transacciones",
            index=False,
        )

        # =========================
        # FORMATO
        # =========================

        workbook = writer.book

        format_workbook(workbook)

    output.seek(0)

    return output.getvalue()


def format_workbook(
    workbook,
) -> None:

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    section_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    white_font = Font(
        color="FFFFFF",
        bold=True,
    )

    bold_font = Font(
        bold=True,
    )

    # =========================
    # TODAS LAS HOJAS
    # =========================

    for worksheet in workbook.worksheets:

        worksheet.freeze_panes = "A2"

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(
                horizontal="center"
            )

        for column_cells in worksheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                value = cell.value

                if value is not None:
                    max_length = max(
                        max_length,
                        len(str(value)),
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                50,
            )

    # =========================
    # RESUMEN
    # =========================

    summary_sheet = workbook["Resumen"]

    section_titles = {
        "INDICADORES FINANCIEROS",
        "INDICADORES COMERCIALES",
        "CENDRA SCORE",
        "RIESGO DE CONCENTRACIÓN",
    }

    for row in summary_sheet.iter_rows():

        first_cell = row[0]

        if first_cell.value in section_titles:

            for cell in row:
                cell.fill = section_fill
                cell.font = bold_font

    # Formato monetario
    currency_labels = {
        "Ventas",
        "Gastos",
        "Utilidad",
        "Ticket promedio",
    }

    # Formato porcentaje
    percentage_labels = {
        "Margen",
        "Cobertura del análisis",
        "Participación cliente",
        "Participación producto",
    }

    for row in summary_sheet.iter_rows(
        min_row=2
    ):

        label = row[0].value
        value_cell = row[1]

        if label in currency_labels:
            value_cell.number_format = (
                '"S/ " #,##0.00'
            )

        if label in percentage_labels:
            value_cell.number_format = (
                "0.0%"
            )